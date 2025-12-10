from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
import io
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import barcode
from barcode.writer import ImageWriter
import qrcode
from PIL import Image, ImageDraw, ImageFont

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# File upload directories
UPLOAD_DIR = ROOT_DIR / 'uploads'
CERTIFICATES_DIR = UPLOAD_DIR / 'certificates'
MANUALS_DIR = UPLOAD_DIR / 'manuals'
RECEIPTS_DIR = UPLOAD_DIR / 'receipts'
CERTIFICATES_DIR.mkdir(parents=True, exist_ok=True)
MANUALS_DIR.mkdir(parents=True, exist_ok=True)
RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password_hash: str
    role: str  # 'admin' or 'viewer'
    full_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    role: str
    full_name: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

class Tool(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    equipment_name: str
    brand_type: str
    serial_no: str
    inventory_code: str
    periodic_inspection_date: Optional[str] = None
    calibration_date: Optional[str] = None
    calibration_validity_months: int = 12  # Default 12 months
    condition: str  # 'Good' or 'Damaged'
    description: Optional[str] = None
    equipment_location: str
    calibration_certificate: Optional[str] = None  # File path
    equipment_manual: Optional[str] = None  # File path
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ToolCreate(BaseModel):
    equipment_name: str
    brand_type: str
    serial_no: str
    inventory_code: str
    periodic_inspection_date: Optional[str] = None
    calibration_date: Optional[str] = None
    calibration_validity_months: int = 12
    condition: str
    description: Optional[str] = None
    equipment_location: str

class ToolResponse(BaseModel):
    id: str
    equipment_name: str
    brand_type: str
    serial_no: str
    inventory_code: str
    periodic_inspection_date: Optional[str]
    calibration_date: Optional[str]
    calibration_validity_months: int
    calibration_expiry_date: Optional[str]
    status: str  # 'Valid', 'Expired', 'Expiring Soon'
    condition: str
    description: Optional[str]
    equipment_location: str
    calibration_certificate: Optional[str]
    equipment_manual: Optional[str]

class LoanEquipment(BaseModel):
    equipment_name: str
    serial_no: str
    condition: str

class Loan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    borrower_name: str
    loan_date: str
    return_date: str
    equipments: List[LoanEquipment]  # Max 5 items
    project_name: str
    wbs_project_no: str
    project_location: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class LoanCreate(BaseModel):
    borrower_name: str
    loan_date: str
    return_date: str
    equipments: List[LoanEquipment]
    project_name: str
    wbs_project_no: str
    project_location: str

class Calibration(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    device_name: str
    serial_no: str
    calibration_date: str
    calibration_expiry_date: str
    device_condition: str
    calibration_agency: str
    calibration_location: str
    person_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class CalibrationCreate(BaseModel):
    device_name: str
    serial_no: str
    calibration_date: str
    calibration_expiry_date: str
    device_condition: str
    calibration_agency: str
    calibration_location: str
    person_name: str

# Stock Management Models
class StockItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    item_name: str
    brand_specifications: str
    available_quantity: int
    unit: str
    description: Optional[str] = None
    purchase_receipt: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StockItemCreate(BaseModel):
    item_name: str
    brand_specifications: str
    available_quantity: int
    unit: str
    description: Optional[str] = None

class StockItemUpdate(BaseModel):
    item_name: Optional[str] = None
    brand_specifications: Optional[str] = None
    available_quantity: Optional[int] = None
    unit: Optional[str] = None
    description: Optional[str] = None

class StockConsume(BaseModel):
    item_id: str
    quantity: int
    reason: Optional[str] = None

# Helper functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=7)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Could not validate credentials")
    
    user = await db.users.find_one({"username": username}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    return user

async def get_admin_user(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

def calculate_tool_status(calibration_date: Optional[str], validity_months: int):
    if not calibration_date:
        return "Unknown", None
    
    try:
        # Parse calibration date (handle both with and without timezone)
        if 'T' in calibration_date or '+' in calibration_date or 'Z' in calibration_date:
            cal_date = datetime.fromisoformat(calibration_date.replace('Z', '+00:00'))
        else:
            # Date only string - treat as UTC
            cal_date = datetime.fromisoformat(calibration_date).replace(tzinfo=timezone.utc)
        
        expiry_date = cal_date + timedelta(days=validity_months * 30)
        expiry_str = expiry_date.strftime('%Y-%m-%d')
        
        now = datetime.now(timezone.utc)
        days_until_expiry = (expiry_date - now).days
        
        if days_until_expiry < 0:
            return "Expired", expiry_str
        elif days_until_expiry <= 90:  # 3 months
            return "Expiring Soon", expiry_str
        else:
            return "Valid", expiry_str
    except Exception as e:
        return "Unknown", None

# Initialize default admin user
@app.on_event("startup")
async def startup_db():
    # Create default admin if not exists
    admin_exists = await db.users.find_one({"username": "admin"})
    if not admin_exists:
        admin = User(
            username="admin",
            password_hash=hash_password("admin123"),
            role="admin",
            full_name="System Administrator"
        )
        doc = admin.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        logger.info("Default admin user created: username=admin, password=admin123")

# Auth endpoints
@api_router.post("/auth/login", response_model=TokenResponse)
async def login(user_login: UserLogin):
    user = await db.users.find_one({"username": user_login.username}, {"_id": 0})
    if not user or not verify_password(user_login.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    access_token = create_access_token(data={"sub": user["username"]})
    user_response = UserResponse(
        id=user["id"],
        username=user["username"],
        role=user["role"],
        full_name=user["full_name"]
    )
    return TokenResponse(access_token=access_token, token_type="bearer", user=user_response)

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        id=current_user["id"],
        username=current_user["username"],
        role=current_user["role"],
        full_name=current_user["full_name"]
    )

# Tool endpoints
@api_router.get("/tools", response_model=List[ToolResponse])
async def get_tools(current_user: dict = Depends(get_current_user)):
    tools = await db.tools.find({}, {"_id": 0}).to_list(1000)
    
    response = []
    for tool in tools:
        status, expiry_date = calculate_tool_status(
            tool.get('calibration_date'),
            tool.get('calibration_validity_months', 12)
        )
        response.append(ToolResponse(
            id=tool['id'],
            equipment_name=tool['equipment_name'],
            brand_type=tool['brand_type'],
            serial_no=tool['serial_no'],
            inventory_code=tool['inventory_code'],
            periodic_inspection_date=tool.get('periodic_inspection_date'),
            calibration_date=tool.get('calibration_date'),
            calibration_validity_months=tool.get('calibration_validity_months', 12),
            calibration_expiry_date=expiry_date,
            status=status,
            condition=tool['condition'],
            description=tool.get('description'),
            equipment_location=tool['equipment_location'],
            calibration_certificate=tool.get('calibration_certificate', None),
            equipment_manual=tool.get('equipment_manual', None)
        ))
    
    return response

@api_router.post("/tools", response_model=ToolResponse)
async def create_tool(tool_create: ToolCreate, current_user: dict = Depends(get_admin_user)):
    tool = Tool(**tool_create.model_dump())
    doc = tool.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.tools.insert_one(doc)
    
    status, expiry_date = calculate_tool_status(
        tool.calibration_date,
        tool.calibration_validity_months
    )
    
    return ToolResponse(
        id=tool.id,
        equipment_name=tool.equipment_name,
        brand_type=tool.brand_type,
        serial_no=tool.serial_no,
        inventory_code=tool.inventory_code,
        periodic_inspection_date=tool.periodic_inspection_date,
        calibration_date=tool.calibration_date,
        calibration_validity_months=tool.calibration_validity_months,
        calibration_expiry_date=expiry_date,
        status=status,
        condition=tool.condition,
        description=tool.description,
        equipment_location=tool.equipment_location,
        calibration_certificate=tool.calibration_certificate,
        equipment_manual=tool.equipment_manual
    )

@api_router.put("/tools/{tool_id}", response_model=ToolResponse)
async def update_tool(tool_id: str, tool_update: ToolCreate, current_user: dict = Depends(get_admin_user)):
    existing_tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not existing_tool:
        raise HTTPException(status_code=404, detail="Tool not found")
    
    update_data = tool_update.model_dump()
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.tools.update_one({"id": tool_id}, {"$set": update_data})
    
    status, expiry_date = calculate_tool_status(
        update_data.get('calibration_date'),
        update_data.get('calibration_validity_months', 12)
    )
    
    return ToolResponse(
        id=tool_id,
        equipment_name=update_data['equipment_name'],
        brand_type=update_data['brand_type'],
        serial_no=update_data['serial_no'],
        inventory_code=update_data['inventory_code'],
        periodic_inspection_date=update_data.get('periodic_inspection_date'),
        calibration_date=update_data.get('calibration_date'),
        calibration_validity_months=update_data.get('calibration_validity_months', 12),
        calibration_expiry_date=expiry_date,
        status=status,
        condition=update_data['condition'],
        description=update_data.get('description'),
        equipment_location=update_data['equipment_location']
    )

@api_router.delete("/tools/{tool_id}")
async def delete_tool(tool_id: str, current_user: dict = Depends(get_admin_user)):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if tool:
        # Delete associated files
        if tool.get('calibration_certificate'):
            cert_path = ROOT_DIR / tool['calibration_certificate']
            if cert_path.exists():
                cert_path.unlink()
        if tool.get('equipment_manual'):
            manual_path = ROOT_DIR / tool['equipment_manual']
            if manual_path.exists():
                manual_path.unlink()
    
    result = await db.tools.delete_one({"id": tool_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tool not found")
    return {"message": "Tool deleted successfully"}

@api_router.post("/tools/{tool_id}/upload-certificate")
async def upload_certificate(
    tool_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_admin_user)
):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(status_code=404, detail="Tool not found")
    
    # Save file
    file_extension = Path(file.filename).suffix
    filename = f"{tool_id}_certificate{file_extension}"
    file_path = CERTIFICATES_DIR / filename
    
    with file_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Update database
    relative_path = f"uploads/certificates/{filename}"
    await db.tools.update_one(
        {"id": tool_id},
        {"$set": {"calibration_certificate": relative_path, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Certificate uploaded successfully", "file_path": relative_path}

@api_router.post("/tools/{tool_id}/upload-manual")
async def upload_manual(
    tool_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_admin_user)
):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(status_code=404, detail="Tool not found")
    
    # Save file
    file_extension = Path(file.filename).suffix
    filename = f"{tool_id}_manual{file_extension}"
    file_path = MANUALS_DIR / filename
    
    with file_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Update database
    relative_path = f"uploads/manuals/{filename}"
    await db.tools.update_one(
        {"id": tool_id},
        {"$set": {"equipment_manual": relative_path, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Manual uploaded successfully", "file_path": relative_path}

@api_router.get("/tools/{tool_id}/download-certificate")
async def download_certificate(tool_id: str, current_user: dict = Depends(get_current_user)):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool or not tool.get('calibration_certificate'):
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    file_path = ROOT_DIR / tool['calibration_certificate']
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(file_path, filename=f"{tool['equipment_name']}_certificate{file_path.suffix}")

@api_router.get("/tools/{tool_id}/download-manual")
async def download_manual(tool_id: str, current_user: dict = Depends(get_current_user)):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool or not tool.get('equipment_manual'):
        raise HTTPException(status_code=404, detail="Manual not found")
    
    file_path = ROOT_DIR / tool['equipment_manual']
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(file_path, filename=f"{tool['equipment_name']}_manual{file_path.suffix}")

@api_router.get("/tools/{tool_id}/barcode")
async def generate_barcode(tool_id: str, current_user: dict = Depends(get_current_user)):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(status_code=404, detail="Tool not found")
    
    # Calculate status
    status, expiry_date = calculate_tool_status(
        tool.get('calibration_date'),
        tool.get('calibration_validity_months', 12)
    )
    
    # Create QR code data
    qr_data = f"""Equipment Information:
Device Name: {tool['equipment_name']}
Serial Number: {tool['serial_no']}
Owner: PT Biro Klasifikasi Indonesia
Calibration Expiry: {expiry_date or 'N/A'}
Status: {status}"""
    
    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    
    # Create QR code image
    qr_img = qr.make_image(fill_color="black", back_color="white")
    
    # Create final image with QR code and text
    img_width, img_height = 800, 650
    img = Image.new('RGB', (img_width, img_height), 'white')
    draw = ImageDraw.Draw(img)
    
    # Resize and center QR code
    qr_img = qr_img.resize((350, 350))
    qr_position = ((img_width - 350) // 2, 50)
    img.paste(qr_img, qr_position)
    
    # Add text information below QR code
    try:
        font_large = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 22)
        font_medium = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 16)
        font_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14)
    except:
        font_large = ImageFont.load_default()
        font_medium = ImageFont.load_default()
        font_small = ImageFont.load_default()
    
    # Draw text information
    y_pos = 420
    
    # Equipment name (centered)
    text = f"{tool['equipment_name']}"
    bbox = draw.textbbox((0, 0), text, font=font_large)
    text_width = bbox[2] - bbox[0]
    x_centered = (img_width - text_width) // 2
    draw.text((x_centered, y_pos), text, fill='black', font=font_large)
    y_pos += 40
    
    # Serial number
    text = f"Serial No: {tool['serial_no']}"
    bbox = draw.textbbox((0, 0), text, font=font_medium)
    text_width = bbox[2] - bbox[0]
    x_centered = (img_width - text_width) // 2
    draw.text((x_centered, y_pos), text, fill='black', font=font_medium)
    y_pos += 35
    
    # Expiry date
    expiry_color = 'red' if status == 'Expired' else 'green' if status == 'Valid' else 'orange'
    text = f"Expiry: {expiry_date or 'N/A'}"
    bbox = draw.textbbox((0, 0), text, font=font_medium)
    text_width = bbox[2] - bbox[0]
    x_centered = (img_width - text_width) // 2
    draw.text((x_centered, y_pos), text, fill=expiry_color, font=font_medium)
    y_pos += 35
    
    # Status
    text = f"Status: {status}"
    bbox = draw.textbbox((0, 0), text, font=font_medium)
    text_width = bbox[2] - bbox[0]
    x_centered = (img_width - text_width) // 2
    draw.text((x_centered, y_pos), text, fill=expiry_color, font=font_medium)
    y_pos += 45
    
    # Owner (centered, blue)
    text = "PT Biro Klasifikasi Indonesia"
    bbox = draw.textbbox((0, 0), text, font=font_large)
    text_width = bbox[2] - bbox[0]
    x_centered = (img_width - text_width) // 2
    draw.text((x_centered, y_pos), text, fill='blue', font=font_large)
    
    # Save to buffer
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    return StreamingResponse(buffer, media_type="image/png", headers={
        "Content-Disposition": f"attachment; filename=qrcode_{tool['serial_no']}.png"
    })

@api_router.get("/tools/export/excel")
async def export_tools_excel(current_user: dict = Depends(get_current_user)):
    tools = await db.tools.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tool Status"
    
    # Headers
    headers = [
        "No.", "Equipment Name", "Brand/Type", "Serial No.", "Inventory Code",
        "Periodic Inspection Date", "Calibration Date", "Calibration Expiry Date",
        "Status", "Condition", "Description", "Equipment Location"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, tool in enumerate(tools, 2):
        status, expiry_date = calculate_tool_status(
            tool.get('calibration_date'),
            tool.get('calibration_validity_months', 12)
        )
        
        row_data = [
            row_num - 1,
            tool['equipment_name'],
            tool['brand_type'],
            tool['serial_no'],
            tool['inventory_code'],
            tool.get('periodic_inspection_date', ''),
            tool.get('calibration_date', ''),
            expiry_date or '',
            status,
            tool['condition'],
            tool.get('description', ''),
            tool['equipment_location']
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Adjust column widths
    column_widths = [5, 25, 20, 15, 15, 20, 18, 20, 15, 12, 30, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tool_status.xlsx"}
    )

# Loan endpoints
@api_router.get("/loans", response_model=List[Loan])
async def get_loans(current_user: dict = Depends(get_current_user)):
    loans = await db.loans.find({}, {"_id": 0}).to_list(1000)
    return loans

@api_router.post("/loans", response_model=Loan)
async def create_loan(loan_create: LoanCreate, current_user: dict = Depends(get_admin_user)):
    if len(loan_create.equipments) > 5:
        raise HTTPException(status_code=400, detail="Maximum 5 equipments allowed per loan")
    
    loan = Loan(**loan_create.model_dump(), created_by=current_user["username"])
    doc = loan.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.loans.insert_one(doc)
    return loan

@api_router.get("/loans/{loan_id}/pdf")
async def get_loan_pdf(loan_id: str, current_user: dict = Depends(get_current_user)):
    loan = await db.loans.find_one({"id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    # Create PDF
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "EQUIPMENT LOAN FORM")
    
    # Form details
    p.setFont("Helvetica", 10)
    y = height - 100
    
    p.drawString(50, y, f"Borrower Name: {loan['borrower_name']}")
    y -= 20
    p.drawString(50, y, f"Loan Date: {loan['loan_date']}")
    y -= 20
    p.drawString(50, y, f"Return Date: {loan['return_date']}")
    y -= 20
    p.drawString(50, y, f"Project Name: {loan['project_name']}")
    y -= 20
    p.drawString(50, y, f"WBS Project No.: {loan['wbs_project_no']}")
    y -= 20
    p.drawString(50, y, f"Project Location: {loan['project_location']}")
    y -= 30
    
    # Equipment table
    p.setFont("Helvetica-Bold", 11)
    p.drawString(50, y, "Equipment Details:")
    y -= 20
    
    # Table headers
    p.setFont("Helvetica-Bold", 9)
    p.drawString(50, y, "No.")
    p.drawString(100, y, "Equipment Name")
    p.drawString(300, y, "Serial No.")
    p.drawString(450, y, "Condition")
    y -= 15
    
    # Equipment rows
    p.setFont("Helvetica", 9)
    for idx, eq in enumerate(loan['equipments'], 1):
        p.drawString(50, y, str(idx))
        p.drawString(100, y, eq['equipment_name'][:30])
        p.drawString(300, y, eq['serial_no'])
        p.drawString(450, y, eq['condition'])
        y -= 15
    
    y -= 30
    
    # Signature section
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "SIGNATURES")
    y -= 30
    
    signatures = [
        ("Borrower", 50),
        ("Submitter", 180),
        ("Approver", 310),
        ("Coordinator", 440)
    ]
    
    for sig_label, x_pos in signatures:
        p.setFont("Helvetica", 9)
        p.drawString(x_pos, y, sig_label + ":")
        p.line(x_pos, y - 5, x_pos + 100, y - 5)
        p.drawString(x_pos, y - 20, "Name:")
        p.line(x_pos + 35, y - 25, x_pos + 100, y - 25)
        p.drawString(x_pos, y - 35, "Date:")
        p.line(x_pos + 30, y - 40, x_pos + 100, y - 40)
    
    y -= 80
    p.setFont("Helvetica", 9)
    p.drawString(50, y, "Senior Manager:")
    p.line(150, y - 5, 300, y - 5)
    p.drawString(50, y - 20, "Name:")
    p.line(100, y - 25, 300, y - 25)
    p.drawString(50, y - 35, "Date:")
    p.line(100, y - 40, 300, y - 40)
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=loan_{loan_id}.pdf"}
    )

# Calibration endpoints
@api_router.get("/calibrations", response_model=List[Calibration])
async def get_calibrations(current_user: dict = Depends(get_current_user)):
    calibrations = await db.calibrations.find({}, {"_id": 0}).to_list(1000)
    return calibrations

@api_router.post("/calibrations", response_model=Calibration)
async def create_calibration(cal_create: CalibrationCreate, current_user: dict = Depends(get_admin_user)):
    calibration = Calibration(**cal_create.model_dump(), created_by=current_user["username"])
    doc = calibration.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.calibrations.insert_one(doc)
    
    # Update tool calibration if exists
    await db.tools.update_one(
        {"serial_no": cal_create.serial_no},
        {"$set": {
            "calibration_date": cal_create.calibration_date,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return calibration

# Stock Management endpoints
@api_router.get("/stock", response_model=List[StockItem])
async def get_stock_items(current_user: dict = Depends(get_current_user)):
    items = await db.stock_items.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/stock", response_model=StockItem)
async def create_stock_item(item_create: StockItemCreate, current_user: dict = Depends(get_admin_user)):
    item = StockItem(**item_create.model_dump())
    doc = item.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.stock_items.insert_one(doc)
    return item

@api_router.put("/stock/{item_id}", response_model=StockItem)
async def update_stock_item(
    item_id: str, 
    item_update: StockItemUpdate, 
    current_user: dict = Depends(get_admin_user)
):
    existing_item = await db.stock_items.find_one({"id": item_id}, {"_id": 0})
    if not existing_item:
        raise HTTPException(status_code=404, detail="Stock item not found")
    
    # If quantity is being updated, add to existing quantity (stock addition)
    update_data = item_update.model_dump(exclude_unset=True)
    if 'available_quantity' in update_data:
        update_data['available_quantity'] = existing_item['available_quantity'] + update_data['available_quantity']
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.stock_items.update_one({"id": item_id}, {"$set": update_data})
    
    updated_item = await db.stock_items.find_one({"id": item_id}, {"_id": 0})
    return StockItem(**updated_item)

@api_router.delete("/stock/{item_id}")
async def delete_stock_item(item_id: str, current_user: dict = Depends(get_admin_user)):
    item = await db.stock_items.find_one({"id": item_id}, {"_id": 0})
    if item:
        # Delete associated receipt file
        if item.get('purchase_receipt'):
            receipt_path = ROOT_DIR / item['purchase_receipt']
            if receipt_path.exists():
                receipt_path.unlink()
    
    result = await db.stock_items.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stock item not found")
    return {"message": "Stock item deleted successfully"}

@api_router.post("/stock/consume")
async def consume_stock(consume: StockConsume, current_user: dict = Depends(get_current_user)):
    """Reduce stock quantity when consuming items"""
    item = await db.stock_items.find_one({"id": consume.item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Stock item not found")
    
    if item['available_quantity'] < consume.quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient stock. Available: {item['available_quantity']} {item['unit']}"
        )
    
    new_quantity = item['available_quantity'] - consume.quantity
    
    await db.stock_items.update_one(
        {"id": consume.item_id},
        {"$set": {
            "available_quantity": new_quantity,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": "Stock consumed successfully",
        "item_name": item['item_name'],
        "consumed_quantity": consume.quantity,
        "remaining_quantity": new_quantity,
        "unit": item['unit']
    }

@api_router.post("/stock/{item_id}/upload-receipt")
async def upload_receipt(
    item_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_admin_user)
):
    item = await db.stock_items.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Stock item not found")
    
    # Save file
    file_extension = Path(file.filename).suffix
    filename = f"{item_id}_receipt{file_extension}"
    file_path = RECEIPTS_DIR / filename
    
    with file_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Update database
    relative_path = f"uploads/receipts/{filename}"
    await db.stock_items.update_one(
        {"id": item_id},
        {"$set": {"purchase_receipt": relative_path, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Receipt uploaded successfully", "file_path": relative_path}

@api_router.get("/stock/{item_id}/download-receipt")
async def download_receipt(item_id: str, current_user: dict = Depends(get_current_user)):
    item = await db.stock_items.find_one({"id": item_id}, {"_id": 0})
    if not item or not item.get('purchase_receipt'):
        raise HTTPException(status_code=404, detail="Receipt not found")
    
    file_path = ROOT_DIR / item['purchase_receipt']
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(file_path, filename=f"{item['item_name']}_receipt{file_path.suffix}")

# Analysis endpoints
@api_router.get("/analysis/tools-usage")
async def get_tools_usage_analysis(current_user: dict = Depends(get_current_user)):
    """Analyze which tools are frequently used based on loan records"""
    loans = await db.loans.find({}, {"_id": 0}).to_list(1000)
    
    tool_usage = {}
    for loan in loans:
        for equipment in loan.get('equipments', []):
            name = equipment['equipment_name']
            if name in tool_usage:
                tool_usage[name] += 1
            else:
                tool_usage[name] = 1
    
    # Sort by frequency
    sorted_usage = sorted(tool_usage.items(), key=lambda x: x[1], reverse=True)
    return [{"equipment_name": name, "usage_count": count} for name, count in sorted_usage[:10]]

@api_router.get("/analysis/tools-damaged")
async def get_tools_damaged_analysis(current_user: dict = Depends(get_current_user)):
    """Analyze damaged tools by type and brand"""
    tools = await db.tools.find({"condition": "Damaged"}, {"_id": 0}).to_list(1000)
    
    by_type = {}
    by_brand = {}
    
    for tool in tools:
        # Count by equipment name (type)
        name = tool['equipment_name']
        if name in by_type:
            by_type[name] += 1
        else:
            by_type[name] = 1
        
        # Count by brand
        brand = tool['brand_type']
        if brand in by_brand:
            by_brand[brand] += 1
        else:
            by_brand[brand] = 1
    
    return {
        "by_type": [{"type": k, "count": v} for k, v in sorted(by_type.items(), key=lambda x: x[1], reverse=True)],
        "by_brand": [{"brand": k, "count": v} for k, v in sorted(by_brand.items(), key=lambda x: x[1], reverse=True)],
        "total_damaged": len(tools)
    }

@api_router.get("/analysis/tools-lost")
async def get_tools_lost_analysis(current_user: dict = Depends(get_current_user)):
    """Analyze lost tools - tools with status Unknown or never returned from loans"""
    # For now, we'll identify potentially lost tools as those with Unknown status
    tools = await db.tools.find({}, {"_id": 0}).to_list(1000)
    
    lost_candidates = []
    for tool in tools:
        status, _ = calculate_tool_status(
            tool.get('calibration_date'),
            tool.get('calibration_validity_months', 12)
        )
        if status == "Unknown":
            lost_candidates.append({
                "equipment_name": tool['equipment_name'],
                "serial_no": tool['serial_no'],
                "brand_type": tool['brand_type'],
                "location": tool['equipment_location']
            })
    
    return {
        "potential_lost": lost_candidates,
        "total": len(lost_candidates)
    }

@api_router.get("/analysis/stock-requested")
async def get_stock_requested_analysis(current_user: dict = Depends(get_current_user)):
    """Analyze frequently requested stock items based on low quantities"""
    stock_items = await db.stock_items.find({}, {"_id": 0}).to_list(1000)
    
    # Items with low stock are frequently requested
    low_stock_items = [
        {
            "item_name": item['item_name'],
            "brand_specifications": item['brand_specifications'],
            "available_quantity": item['available_quantity'],
            "unit": item['unit']
        }
        for item in stock_items
        if item['available_quantity'] < 50
    ]
    
    # Sort by lowest quantity (most requested)
    low_stock_items.sort(key=lambda x: x['available_quantity'])
    
    return {
        "frequently_requested": low_stock_items,
        "total_low_stock": len(low_stock_items)
    }

@api_router.get("/analysis/stock-purchased")
async def get_stock_purchased_analysis(current_user: dict = Depends(get_current_user)):
    """Analyze frequently purchased items by brand and name"""
    stock_items = await db.stock_items.find({}, {"_id": 0}).to_list(1000)
    
    by_brand = {}
    by_item = {}
    
    for item in stock_items:
        # Count by brand
        brand = item['brand_specifications']
        if brand in by_brand:
            by_brand[brand] += 1
        else:
            by_brand[brand] = 1
        
        # Count by item name
        name = item['item_name']
        if name in by_item:
            by_item[name] += 1
        else:
            by_item[name] = 1
    
    return {
        "by_brand": [{"brand": k, "count": v} for k, v in sorted(by_brand.items(), key=lambda x: x[1], reverse=True)[:10]],
        "by_item": [{"item_name": k, "count": v} for k, v in sorted(by_item.items(), key=lambda x: x[1], reverse=True)[:10]],
        "total_items": len(stock_items)
    }

@api_router.get("/analysis/summary")
async def get_analysis_summary(current_user: dict = Depends(get_current_user)):
    """Get overall summary statistics for analysis dashboard"""
    tools = await db.tools.find({}, {"_id": 0}).to_list(1000)
    loans = await db.loans.find({}, {"_id": 0}).to_list(1000)
    stock_items = await db.stock_items.find({}, {"_id": 0}).to_list(1000)
    
    # Calculate statistics
    damaged_tools = len([t for t in tools if t['condition'] == 'Damaged'])
    good_tools = len([t for t in tools if t['condition'] == 'Good'])
    total_loans = len(loans)
    low_stock = len([s for s in stock_items if s['available_quantity'] < 50])
    
    return {
        "total_tools": len(tools),
        "damaged_tools": damaged_tools,
        "good_tools": good_tools,
        "damage_rate": round((damaged_tools / len(tools) * 100) if len(tools) > 0 else 0, 1),
        "total_loans": total_loans,
        "total_stock_items": len(stock_items),
        "low_stock_items": low_stock,
        "low_stock_rate": round((low_stock / len(stock_items) * 100) if len(stock_items) > 0 else 0, 1)
    }

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
